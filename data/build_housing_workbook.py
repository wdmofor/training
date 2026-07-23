#!/usr/bin/env python3
"""Build the Week 3 Excel evidence workbook for the African housing panel."""

from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
PANEL = ROOT / "africa_housing_services_panel.csv"
AUDIT = json.loads((ROOT / "dataset_audit.json").read_text(encoding="utf-8"))
OUTPUT = ROOT / "week3_africa_housing_services_data.xlsx"


def write_table(sheet, headers, rows):
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 45)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = width


with PANEL.open(encoding="utf-8", newline="") as handle:
    panel_rows = list(csv.DictReader(handle))

workbook = Workbook()
readme = workbook.active
readme.title = "README"
write_table(readme, ["Item", "Value"], [
    ["Project", "Predicting and explaining housing basic-services deprivation across Africa"],
    ["Unit of analysis", "African country-year"],
    ["Source", "World Bank World Development Indicators API"],
    ["Source URL", "https://data.worldbank.org/"],
    ["License", "CC BY-4.0 as displayed on World Bank indicator pages"],
    ["Rows in panel", AUDIT["rows_written"]],
    ["Complete target rows", AUDIT["target_complete_rows"]],
    ["Complete modeling rows", AUDIT["complete_model_rows"]],
    ["Important limitation", AUDIT["target_caveat"]],
])

source = workbook.create_sheet("source_inventory")
write_table(source, ["Field", "Indicator code", "Source URL", "Definition", "Role"], [
    ["water_access_pct", "SH.H2O.BASW.ZS", "https://data.worldbank.org/indicator/SH.H2O.BASW.ZS", "People using at least basic drinking-water services (%)", "Target component"],
    ["sanitation_access_pct", "SH.STA.BASS.ZS", "https://data.worldbank.org/indicator/SH.STA.BASS.ZS", "People using at least basic sanitation services (%)", "Target component"],
    ["electricity_access_pct", "EG.ELC.ACCS.ZS", "https://data.worldbank.org/indicator/EG.ELC.ACCS.ZS", "Access to electricity (% of population)", "Target component"],
    ["gdp_per_capita_constant_usd", "NY.GDP.PCAP.KD", "https://data.worldbank.org/indicator/NY.GDP.PCAP.KD", "GDP per capita, constant 2015 US$", "Predictor"],
    ["urban_population_pct", "SP.URB.TOTL.IN.ZS", "https://data.worldbank.org/indicator/SP.URB.TOTL.IN.ZS", "Urban population (% of total population)", "Predictor"],
    ["population_density", "EN.POP.DNST", "https://data.worldbank.org/indicator/EN.POP.DNST", "People per square kilometer of land area", "Predictor"],
    ["population_total", "SP.POP.TOTL", "https://data.worldbank.org/indicator/SP.POP.TOTL", "Total population", "Predictor"],
])

dictionary = workbook.create_sheet("data_dictionary")
write_table(dictionary, ["Variable", "Type", "Role", "Timing", "Definition", "Missingness rule"], [
    ["country_code", "String", "Identifier/group", "Country-year", "ISO-3 country code", "Must be present"],
    ["country_name", "String", "Label", "Country-year", "World Bank country name", "Must be present"],
    ["year", "Integer", "Time index", "Country-year", "Calendar year", "2000-2024 only"],
    ["housing_service_deprivation_pct", "Float", "Primary target", "Same year", "Mean of 100 minus water, sanitation, electricity access", "Complete component rows only"],
    ["high_deprivation", "Binary derived", "Sensitivity target", "One-year ahead", "At/above training-fold median", "Threshold learned inside training fold"],
    ["gdp_per_capita_constant_usd", "Float", "Predictor", "Prior year for RQ4", "Constant 2015 US dollars per capita", "Complete-case sensitivity"],
    ["urban_population_pct", "Float", "Predictor", "Prior year for RQ4", "Urban population percentage", "Complete-case sensitivity"],
    ["population_density", "Float", "Predictor", "Prior year for RQ4", "People per square kilometer", "Complete-case sensitivity"],
    ["population_total", "Float", "Predictor", "Prior year for RQ4", "Total population", "Log-transform and complete-case sensitivity"],
])

quality = workbook.create_sheet("quality_checks")
missing_rows = []
for field in ["water_access_pct", "sanitation_access_pct", "electricity_access_pct", "gdp_per_capita_constant_usd", "urban_population_pct", "population_density", "population_total"]:
    missing_rows.append([field, sum(row.get(field, "") == "" for row in panel_rows), len(panel_rows)])
write_table(quality, ["Field", "Missing rows", "Total rows"], missing_rows)

sample = workbook.create_sheet("sample_size")
write_table(sample, ["RQ", "Method", "Assumptions", "Minimum N", "Available comparison"], [
    ["RQ1", "Mean-change precision", "95% CI; SD=10; margin=2", 97, 1189],
    ["RQ2", "Regression planning sensitivity", "5 predictors; medium f2=.15; alpha=.05; power=.80; 20/predictor", 100, 1189],
    ["RQ3", "Omnibus regional comparison", "5 subregions; Cohen f=.25; alpha=.05; power=.80", 196, 1189],
    ["RQ4", "Two-proportion sensitivity", "alpha=.05; power=.80; meaningful difference=.15; balanced case", 350, 1189],
    ["Final", "Maximum with 10% allowance", "max(97,100,196,350) x 1.10", 385, 1189],
])

mapping = workbook.create_sheet("analysis_mapping")
write_table(mapping, ["RQ", "Estimand", "Method", "Primary output", "Limitation"], [
    ["RQ1", "Mean annual change in deprivation", "Country trajectories and fixed-effects trend", "Trend table and confidence intervals", "Ecological country-year inference"],
    ["RQ2", "Adjusted association with deprivation", "Regression with country/year controls", "Coefficients and robust uncertainty", "Association is not causation"],
    ["RQ3", "Subregional mean difference", "Year-adjusted omnibus comparison", "Regional contrasts", "Repeated countries and unequal coverage"],
    ["RQ4", "One-year-ahead high-deprivation classification", "Chronological baseline and ML comparison", "ROC-AUC, PR-AUC, F1, calibration", "Small number of independent countries"],
])

literature = workbook.create_sheet("literature_matrix")
write_table(literature, [
    "Article", "APA citation", "DOI/URL", "Journal/year", "Geography/unit",
    "Sample", "Variables/target", "Method", "Validation/metrics", "Findings",
    "Limitations", "Similarity", "Difference", "Gap contribution", "Verification",
], [
    ["Gradin (2013)", "Gradín, C. (2013). Race, poverty and deprivation in South Africa. Journal of African Economies, 22(2), 187-238.", "https://doi.org/10.1093/jae/ejs019", "Journal of African Economies (2013)", "South Africa; household survey", "Exact N not reported in accessible abstract", "Poverty/material deprivation; race, education, demographics, residence", "Conditional/counterfactual distribution comparisons", "Not reported in accessible abstract", "Racial deprivation gaps reflected cumulative disadvantage and historical inequality.", "Full text restricted; exact details require confirmation.", "African deprivation; place/development context", "Household, single-country, non-panel; no forecast", "Adds evidence on spatial and historical heterogeneity.", "Publisher abstract + DOI metadata"],
    ["Kamndaya et al. (2014)", "Kamndaya, M., et al. (2014). Material deprivation affects high sexual risk behavior among young people in urban slums, South Africa. Journal of Urban Health, 91, 581-594.", "https://doi.org/10.1007/s11524-013-9856-1", "Journal of Urban Health (2014)", "South African urban slums; individual survey", "Exact N not reported in accessible evidence", "Material deprivation and sexual-risk outcome", "Observational association analysis; exact estimator pending", "Metrics not reported in accessible evidence", "Material deprivation was linked to vulnerability in an urban-slum context.", "Publisher page did not expose full operational details.", "Multidimensional deprivation; African setting", "Individual/urban-slum data; behavioral outcome; not country-year", "Motivates sensitivity to urbanization and service context.", "DOI metadata + publisher record"],
    ["Marutlulle (2021)", "Marutlulle, N. K. (2021). A critical analysis of housing inadequacy in South Africa and its ramifications. Africa's Public Service Delivery & Performance Review, 9(1), a372.", "https://doi.org/10.4102/apsdpr.v9i1.372", "Africa's Public Service Delivery & Performance Review (2021)", "South Africa; literature/document/policy review", "No respondent sample; secondary sources", "Housing inadequacy, informal settlements, migration, population, service consequences", "Exploratory qualitative review", "No statistical validation metrics", "Housing inadequacy was associated with protests, health risks, fires, flooding, violence, corruption, and xenophobia.", "Review/secondary evidence; no causal or predictive estimate.", "Direct housing/basic-service relevance in Africa", "Single-country qualitative policy analysis; not panel or forecast", "Provides mechanisms aggregate indicators cannot capture.", "Full text and publisher abstract"],
    ["Lin & Okyere (2022)", "Lin, B., & Okyere, M. A. (2022). Race and energy poverty: The moderating role of subsidies in South Africa. Energy Economics, 110, 106464.", "https://doi.org/10.1016/j.eneco.2022.106464", "Energy Economics (2022)", "South Africa; household/consumer evidence", "Exact N not visible in accessible evidence", "Energy poverty, race, subsidies, moderation", "Moderation analysis; exact estimator pending full text", "Metrics not visible in accessible evidence", "Studies how subsidies moderate racial differences in energy poverty.", "Full text not accessible; do not infer exact N or coefficients.", "Electricity/energy poverty dimension", "Single-country distributional analysis; no country-year forecast", "Supports treating electricity as one dimension, not total housing adequacy.", "DOI metadata + index evidence"],
    ["Yeh et al. (2020)", "Yeh, C., Perez, A., Driscoll, A., Athey, S., Bailard, C., Blumenstock, J. E., et al. (2020). Using publicly available satellite imagery and deep learning to understand economic well-being in Africa. Nature Communications, 11, 2583.", "https://doi.org/10.1038/s41467-020-16185-w", "Nature Communications (2020)", "Africa; spatial units linked to surveys and imagery", "Exact country/spatial N pending article-table check", "Economic well-being labels and satellite features", "Deep-learning prediction with geographic holdouts", "Held-out geographic metrics; copy exact values from article tables", "Public satellite imagery and deep learning predicted spatial economic-well-being variation.", "Transferability and label quality; not direct housing-services measurement.", "Public, geographically extensive African prediction", "Subnational imagery model, not WDI country-year panel", "Supports geographic blocking and public-data prediction.", "DOI metadata + publisher record"],
    ["de Milliano & Plavgo (2018)", "de Milliano, M., & Plavgo, I. (2018). Analysing multidimensional child poverty in sub-Saharan Africa. Child Indicators Research, 11, 923-945.", "https://doi.org/10.1007/s12187-017-9488-1", "Child Indicators Research (2018)", "Sub-Saharan Africa; child/household survey evidence", "Exact pooled N not reported in accessible evidence", "Multidimensional child poverty, living conditions, services", "Deprivation-index and comparative analysis", "Exact robustness metrics require full text", "Child poverty is multidimensional and context-dependent; monetary poverty misses deprivations.", "Cross-country comparability and indicator availability.", "Regional African multidimensional deprivation framework", "Child/household microdata; irregular surveys; broader poverty construct", "Supports explicit index construction and indicator limitations.", "DOI metadata + abstract evidence"],
    ["Bouzarovski & Petrova (2015)", "Bouzarovski, S., & Petrova, S. (2015). A global perspective on domestic energy deprivation: Overcoming the energy poverty-fuel poverty binary. Energy Research & Social Science, 10, 31-40.", "https://doi.org/10.1016/j.erss.2015.06.007", "Energy Research & Social Science (2015)", "Global conceptual/comparative literature", "Not applicable", "Domestic energy services, vulnerability, socio-technical pathways", "Conceptual framework", "No predictive validation metrics", "Energy deprivation is inability to attain socially necessary domestic energy services, not only fuel/income poverty.", "Conceptual; does not estimate the active target.", "Informs electricity-component interpretation", "Conceptual, not empirical country-year model", "Prevents equating electricity access with complete energy poverty.", "Full abstract + DOI metadata"],
    ["Zimmerman et al. (2022)", "Zimmerman, A., Lund, C., Araya, R., Hessel, P., Sanchez, J., Garman, E., Evans-Lacko, S., Diaz, Y., & Avendano-Pabon, M. (2022). The relationship between multidimensional poverty, income poverty and youth depressive symptoms: Cross-sectional evidence from Mexico, South Africa and Colombia. BMJ Global Health, 7(1), e006960.", "https://doi.org/10.1136/bmjgh-2021-006960", "BMJ Global Health (2022)", "Colombia, Mexico, South Africa; ages 11-25", "Complete-case N=16,173; 1,538 excluded", "14 indicators/five dimensions including housing, sewage, water, overcrowding; depression", "Weighted Poisson regression, country interactions, logistic sensitivity", "IRR, 95% CI, Wald tests, binary sensitivity", "MPI associated with symptoms overall, Mexico, and Colombia, but not South Africa; context mattered.", "Cross-sectional; survey timing differs; adapted indicators; missingness.", "Multidimensional service/housing indicators; African comparison", "Individual cross-section and mental-health outcome, not country-year forecast", "Best bridge to dimension-specific and country-aware analysis.", "Full text HTML/PDF + metadata"],
    ["Agbodji et al. (2015)", "Agbodji, A. E., Batana, Y. M., & Ouédraogo, D. (2015). Gender inequality in multidimensional welfare deprivation in West Africa: The case of Burkina Faso and Togo. International Journal of Social Economics, 42(11), 980-1004.", "https://doi.org/10.1108/IJSE-11-2013-0270", "International Journal of Social Economics (2015)", "Burkina Faso and Togo; household surveys", "Exact analytic N not visible in accessible evidence", "Six dimensions: housing, utilities, assets, education, employment, credit", "Counting approach with gender and regional decomposition", "Decomposition results; no predictive validation metrics", "Women were more deprived; regional disparities existed; contributors differed by country.", "Two countries; cross-sectional surveys; different survey years.", "Housing/utilities, West African geography, regional differences", "Household/gender survey data; no repeated country-year prediction", "Supports regional decomposition and unequal dimensional contributions.", "Publisher abstract + DOI metadata"],
    ["Chiwarawara (2024)", "Chiwarawara, K. (2024). The struggle for housing and basic services in South Africa: A case for service delivery protests. Social Dynamics, 50(2), 148-165.", "https://doi.org/10.1080/02533952.2024.2352193", "Social Dynamics (2024)", "Gugulethu and Khayelitsha, Cape Town", "40 interviews and 4 focus groups overall", "Housing, water, electricity, sanitation, protest demands", "In-depth interviews/focus groups; framing-process analysis", "No statistical validation metrics", "Residents framed protest as a way to accelerate housing/service delivery; housing was a service nexus.", "Localized purposive qualitative evidence; no national prevalence estimate.", "Directly connects housing and the three active service domains", "Local lived experience and mechanisms, not country-year levels", "Supplies context for aggregate patterns and ecological limits.", "Full text HTML + publisher metadata"],
])

panel = workbook.create_sheet("panel_data")
headers = list(panel_rows[0])
write_table(panel, headers, [[row.get(header, "") for header in headers] for row in panel_rows])

workbook.save(OUTPUT)
print(OUTPUT)
